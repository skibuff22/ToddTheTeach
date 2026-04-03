import openpyxl

file_path = r"C:\Users\tdoan\GitHub\Calyx-Ai.net\Project-Tracker.xlsx"

wb = openpyxl.load_workbook(file_path)

print("--- Sprint Plan ---")
ws = wb["Sprint Plan"]
for r in range(1, 20):
    valA = ws.cell(row=r, column=1).value
    valG = ws.cell(row=r, column=7).value
    if valA:
        print(f"Row {r}: Sprint {valA} | Status: {valG}")

print("\n--- Dashboard ---")
wd = wb["Dashboard"]
for r in range(1, 15):
    valA = wd.cell(row=r, column=1).value
    valD = wd.cell(row=r, column=4).value
    if valA:
        print(f"Row {r}: Phase {valA} | Status: {valD}")

