import os
import re
import pandas as pd
import openpyxl
from datetime import datetime
import sys

# Change directory to the root of the website
base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(base_dir, "executions"))
import deploy_now
import deploy_cpanel

def map_backlog_to_sprint(df_backlog):
    sprint_status_map = {}
    total_sprints = 0
    for sprint_val, group in df_backlog.groupby('Sprint'):
        if pd.isna(sprint_val):
            continue
        try:
            sprint = int(float(sprint_val))
            if sprint > total_sprints:
                total_sprints = sprint
        except:
            continue
            
        statuses = group['Status'].astype(str).str.strip().tolist()
        valid_statuses = [s for s in statuses if s and s.lower() != 'nan' and s.lower() != 'none']
        
        if not valid_statuses:
            sprint_status_map[sprint] = 'Not Started'
            continue
            
        all_done = all(s in ['Done', 'Completed'] for s in valid_statuses)
        any_in_progress = any(s in ['In Progress', 'Done', 'Completed'] for s in valid_statuses)
        
        if all_done:
            sprint_status_map[sprint] = 'Completed'
        elif any_in_progress:
            sprint_status_map[sprint] = 'In Progress'
        else:
            sprint_status_map[sprint] = 'Not Started'
            
    return sprint_status_map, total_sprints

def map_backlog_to_phase(df_backlog):
    def extract_phase(epic_str):
        if pd.isna(epic_str):
             return None
        match = re.search(r'Phase\s+(\d+)', str(epic_str), re.IGNORECASE)
        if match:
             return int(match.group(1))
        return None
        
    df_backlog['Phase_Num'] = df_backlog['Epic'].apply(extract_phase)
    
    phase_status_map = {}
    total_phases = 0
    
    for phase_val, group in df_backlog.groupby('Phase_Num'):
        if pd.isna(phase_val):
            continue
        phase = int(phase_val)
        if phase > total_phases:
            total_phases = phase
            
        statuses = group['Status'].astype(str).str.strip().tolist()
        valid_statuses = [s for s in statuses if s and s.lower() != 'nan' and s.lower() != 'none']
        
        if not valid_statuses:
            phase_status_map[phase] = 'Not Started'
            continue
            
        all_done = all(s in ['Done', 'Completed'] for s in valid_statuses)
        any_in_progress = any(s in ['In Progress', 'Done', 'Completed'] for s in valid_statuses)
        
        if all_done:
            phase_status_map[phase] = 'Completed'
        elif any_in_progress:
            phase_status_map[phase] = 'In Progress'
        else:
            phase_status_map[phase] = 'Not Started'
            
    return phase_status_map, total_phases

def update_calyx_progress():
    print(f"[{datetime.now()}] Updating Calyx AI Project Tracker...")
    
    file_path = r"C:\Users\tdoan\GitHub\Calyx-Ai.net\Project-Tracker.xlsx"
    index_html_path = os.path.join(base_dir, "index.html")
    
    if not os.path.exists(file_path):
        print(f"Error: Could not find {file_path}")
        return False

    try:
        # 1. Math computation from Backlog dynamically
        xl = pd.ExcelFile(file_path)
        df_backlog = xl.parse("Backlog")
        
        sprint_status_map, total_sprints = map_backlog_to_sprint(df_backlog)
        phase_status_map, total_phases = map_backlog_to_phase(df_backlog)
        
        print(f"Detected total phases: {total_phases}, total sprints: {total_sprints}")
        print("Computed Sprint Statuses:", sprint_status_map)
        print("Computed Phase Statuses:", phase_status_map)

        # 2. Open Workbook with openpyxl to edit cells losslessly
        wb = openpyxl.load_workbook(file_path)
        
        ws_sprint = wb["Sprint Plan"]
        # Update existing Sprint Plan rows
        existing_sprints = set()
        last_sprint_row = 1
        for r in range(2, 500):
            val_sprint = ws_sprint.cell(row=r, column=1).value
            if val_sprint is None:
                continue
            last_sprint_row = r
            try:
                sprint_num = int(float(val_sprint))
                existing_sprints.add(sprint_num)
                if sprint_num in sprint_status_map:
                    ws_sprint.cell(row=r, column=7).value = sprint_status_map[sprint_num]
            except:
                continue
                
        # Append missing sprints if needed
        next_sprint_row = last_sprint_row + 1
        for s in range(1, total_sprints + 1):
            if s not in existing_sprints:
                ws_sprint.cell(row=next_sprint_row, column=1).value = s
                ws_sprint.cell(row=next_sprint_row, column=7).value = sprint_status_map.get(s, 'Not Started')
                # Attempt to link to a phase from backlog (rough approximation)
                phase_match = df_backlog[df_backlog['Sprint'] == s]['Phase_Num'].dropna().unique()
                if len(phase_match) > 0:
                    ws_sprint.cell(row=next_sprint_row, column=3).value = f"Phase {int(phase_match[0])}"
                next_sprint_row += 1


        # Update Dashboard
        ws_dash = wb["Dashboard"]
        # Also update the "Last Updated" text
        for r in range(1, 4):
            if str(ws_dash.cell(row=r, column=1).value).startswith("Last Updated:"):
                ws_dash.cell(row=r, column=1).value = f"Last Updated: {datetime.now().strftime('%B %d, %Y')}"
                
        existing_phases = set()
        last_phase_row = 5
        for r in range(5, 50):
            val_phase = ws_dash.cell(row=r, column=1).value
            if val_phase is None:
                continue
            last_phase_row = r
            try:
                phase_num = int(float(val_phase))
                existing_phases.add(phase_num)
                if phase_num in phase_status_map:
                    ws_dash.cell(row=r, column=4).value = phase_status_map[phase_num]
            except:
                continue
                
        # Append missing phases into dashboard
        next_phase_row = last_phase_row + 1
        for p in range(1, total_phases + 1):
            if p not in existing_phases:
                ws_dash.cell(row=next_phase_row, column=1).value = p
                # rough name from backlog epic
                name_match = df_backlog[df_backlog['Phase_Num'] == p]['Epic'].dropna().unique()
                name_str = "Unknown Area"
                if len(name_match) > 0:
                    name_str = re.sub(r'(?i)Phase\s*\d+[:\-\s]*', '', name_match[0]).strip()
                ws_dash.cell(row=next_phase_row, column=2).value = name_str
                ws_dash.cell(row=next_phase_row, column=4).value = phase_status_map.get(p, 'Not Started')
                next_phase_row += 1

        # Save workbook safely
        try:
            wb.save(file_path)
            print("Saved updated Project-Tracker.xlsx.")
        except PermissionError:
            print(f"Warning: Could not save to {file_path} because it is currently open in Excel. Website calculation will proceed with computed values.")
        except Exception as e:
            print(f"Warning: Could not save Excel file: {e}")
        
        # 3. Determine explicit active Sprint and Phase for Website
        current_phase = 1
        in_prog_phases = [k for k, v in phase_status_map.items() if v == 'In Progress']
        if in_prog_phases:
            current_phase = max(in_prog_phases)
        else:
            comp_phases = [k for k, v in phase_status_map.items() if v == 'Completed']
            if comp_phases:
                current_phase = min(total_phases, max(comp_phases) + 1)
                
        # Safeguard div 0
        if total_phases == 0: total_phases = 1
        phase_pct = min(100, int((current_phase / total_phases) * 100))

        # Find active sprint
        current_sprint = 1
        in_prog_sprints = [k for k, v in sprint_status_map.items() if v == 'In Progress']
        if in_prog_sprints:
            current_sprint = max(in_prog_sprints)
        else:
            comp_sprints = [k for k, v in sprint_status_map.items() if v == 'Completed']
            if comp_sprints:
                current_sprint = min(total_sprints, max(comp_sprints) + 1)
                
        # Safeguard div 0
        if total_sprints == 0: total_sprints = 1
        sprint_pct = min(100, int((current_sprint / total_sprints) * 100))
        
        print(f"Website values -> Phase: {current_phase}/{total_phases} ({phase_pct}%), Sprint: {current_sprint}/{total_sprints} ({sprint_pct}%)")

        # 4. Modify HTML
        with open(index_html_path, "r", encoding="utf-8") as f:
            html = f.read()

        html = re.sub(
            r'<!-- CALYX_PHASE_TEXT_START -->.*?<!-- CALYX_PHASE_TEXT_END -->',
            f'<!-- CALYX_PHASE_TEXT_START --><span id="calyx-phase-text" style="color: #bbb;">Phase {current_phase} of {total_phases}</span><!-- CALYX_PHASE_TEXT_END -->',
            html,
            flags=re.DOTALL
        )
        
        html = re.sub(
            r'<!-- CALYX_PHASE_BAR_START -->.*?<!-- CALYX_PHASE_BAR_END -->',
            f'<!-- CALYX_PHASE_BAR_START --><div id="calyx-phase-bar" style="height: 100%; width: {phase_pct}%; background: var(--accent); transition: width 1s ease-in-out; border-radius: 3px;"></div><!-- CALYX_PHASE_BAR_END -->',
            html,
            flags=re.DOTALL
        )
        
        html = re.sub(
            r'<!-- CALYX_SPRINT_TEXT_START -->.*?<!-- CALYX_SPRINT_TEXT_END -->',
            f'<!-- CALYX_SPRINT_TEXT_START --><span id="calyx-sprint-text" style="color: #bbb;">Sprint {current_sprint} of {total_sprints}</span><!-- CALYX_SPRINT_TEXT_END -->',
            html,
            flags=re.DOTALL
        )
        
        html = re.sub(
            r'<!-- CALYX_SPRINT_BAR_START -->.*?<!-- CALYX_SPRINT_BAR_END -->',
            f'<!-- CALYX_SPRINT_BAR_START --><div id="calyx-sprint-bar" style="height: 100%; width: {sprint_pct}%; background: linear-gradient(90deg, var(--accent) 0%, rgba(255,255,255,0.8) 100%); transition: width 1s ease-in-out; border-radius: 3px;"></div><!-- CALYX_SPRINT_BAR_END -->',
            html,
            flags=re.DOTALL
        )

        with open(index_html_path, "w", encoding="utf-8") as f:
            f.write(html)
            
        print("Updated index.html successfully.")
        
        # 5. Trigger Deployment
        print("Triggering deployment with deploy_now...")
        os.chdir(base_dir) # ensure we are in the repo root for git commands
        deploy_now.git_push()
        deploy_cpanel.deploy_cpanel()
        
        return True

    except Exception as e:
        print(f"Error during update process: {e}")
        return False

if __name__ == "__main__":
    update_calyx_progress()
